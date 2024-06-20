# main.py
import json
from datetime import datetime
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from helper_functions import (
    calculate_num_leading_space,
    get_hierarchical_string,
    process_table,
)

def processor(file):
    """
    Process and json result as a JSON file in "./parsed_files"

    Args:
    file (str): The name of file in uploads
    """
    excel_file = Path(f"./uploads/{file}")
    simple_ws = get_sheet_from_excel(excel_file)
    # Process the table
    records, datasets_range = process_table(simple_ws)

    parsed_data = parser_to_llm(get_datasets(records, datasets_range))

    output_dir = Path("./parsed_files")

    output_path = output_dir / f"{excel_file.stem}.json"

    # Save the parsed data as a JSON file
    save_as_json(parsed_data, output_path)

    return output_path

def parser_to_llm(datasets):
    """
    Parser the data to be clear for a LLM

    Args:
    datasets (Dict): dataset ranges

    Returns:
    "BALANCE SHEET > Cash & Due from Banks > October 31, 2023 > 70778000",
    "BALANCE SHEET > Short-term Investments > March 31, 2021 > 136000", ...
    """
    parsed_data = []
    for dataset in datasets:
        header_row = list(dataset[0].values())

        hierarchical_levels = {}

        for row in dataset[1:]:
            row_values = list(row.values())

            hierarchical_string = None
            if header_row[0] is None:
                hierarchical_string = get_hierarchical_string(hierarchical_levels, row_values[0])
                # print(f"{hierarchical_string}")
                hierarchical_levels[calculate_num_leading_space(row_values[0])] = (row_values[0])

            for idx, field in enumerate(row_values[1:]):
                if field is None:
                    continue

                if isinstance(header_row[idx + 1], datetime):
                    header_row[idx + 1] = header_row[idx + 1].strftime("%B %d, %Y")

                if hierarchical_string is None:
                    parsed_data.append(f"{header_row[0]} > {row_values[0]} > {header_row[idx+1]} > {field}")
                else:
                    parsed_data.append(f"{hierarchical_string} > {header_row[idx+1]} > {field}")

    return parsed_data

def get_datasets(records, schema):
    datasets = []

    for entry in schema:
        row = entry["row"]
        fields = entry["fields"]
        end_row = entry.get("end_row", len(records))

        dataset = []
        for r in range(row, end_row):
            record = {f"{idx+1}": records[r][idx] for idx in fields}

            # remove None rows
            if any(value is not None for value in record.values()):
                dataset.append(record)

        datasets.append(
            dataset
        )

    # print(datasets)
    return datasets

def get_sheet_from_excel(
    filename: Path, dataOnly: bool = True
) -> Worksheet:
    """
    Get a worksheet from an Excel file

    Args:
    filename (Path): The path to the Excel file
    dataOnly (bool): Whether to load the worksheet with data only (formulas will be evaluated to their values). Default is True.

    Returns:
    Worksheet: The worksheet object
    """
    wb = load_workbook(filename, data_only=dataOnly)
    first_sheet_name = wb.sheetnames[0]  # Get the name of the first sheet
    return wb[first_sheet_name]

def save_as_json(parsed_data: Any, output_path: Path) -> None:
    with output_path.open("w") as json_file:
        json.dump(parsed_data, json_file, indent=4)
